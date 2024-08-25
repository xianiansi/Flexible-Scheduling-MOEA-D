import numpy as np
import matplotlib.pyplot as plt
import mpl_toolkits
import copy
import math
import openpyxl
from scipy.special import comb  # comb 组合数Cmn
from itertools import combinations
from neww import parse
import random
import xlrd
import math
import random
import matplotlib
import plotly as py
from colorama import init, Fore
from matplotlib import colors as mcolors
import time
import plotly.figure_factory as ff
from matplotlib.pyplot import MultipleLocator
from collections import Counter
import matplotlib as mpl
import openpyxl
from mpl_toolkits.mplot3d import Axes3D  # 空间三维画图
from scipy.special import comb  # comb 组合数Cmn
from itertools import combinations
from time import sleep, ctime
import xlrd


# 这是一组合格的随机散点
def generate_point():
    point_list=[]
    n = 50
    x = np.random.uniform(0, 50,size=n) #50*1的数组
    deviate_x = np.random.normal(0, 5, size=n)  #loc,scale
    deviate_y = np.random.normal(0, 5, n)
    deviate_z = np.random.normal(0, 5, n)
    x += deviate_x
    y = x + deviate_y
    z = 0.5*(x+y) + deviate_z

    # **可删**
    print(point_list)
    fig = plt.figure()
    ax3d = mpl_toolkits.mplot3d.Axes3D(fig)
    ax3d.scatter(x, y, z)
    plt.show()
    return (x,y,z)

matplotlib.use('TKAgg')
pyplt = py.offline.plot
color = [
    'lightcoral',
    'coral',
    'darkorange',
    'gold',
    'palegreen',
    'paleturquoise',
    'skyblue',
    'plum',
    'hotpink',
    'pink',
    'tomato',
    'lightgreen',
    'lightskyblue']
newcmp = mpl.colors.LinearSegmentedColormap.from_list(name='chaos', colors=color, N=12)  # N 表示插值后的颜色数目
color_list = [[newcmp(i)[0], newcmp(i)[1], newcmp(i)[2]] for i in range(newcmp.N)]
COLORS = list(color_list)
LEN_COLORS = len(COLORS)
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
init(autoreset=True)

# 工具函数
def decode(C):  # 解码OS
    q = len(C)
    tmp = [0 for i in range(parse.J)]  # 加工到第几个工序
    P = [0 for i in range(q)]
    for i in range(q):
        tmp[int(C[i]) - 1] += 1
        P[i] = C[i] * 100 + tmp[int(C[i]) - 1]
    return np.array(P)  # [101 301 ...] 解码后的染色体个体P （工件和工序）1201
def Write_cell_str(sheet,line, column, string):
    wb = openpyxl.load_workbook("result.xlsx")  # 生成一个已存在的wookbook对象
    wb1 = wb.worksheets[sheet]
    # wb1 = wb.active  # 激活的sheet
    wb1.cell(line, column, string)  # 对应14行B
    wb.save("result.xlsx")  # 保存
def Write_cell(sheet,line, column, targ_i):
    # string = str(array)
    wb = openpyxl.load_workbook("result.xlsx")  # 生成一个已存在的wookbook对象
    # wb1 = wb.active  # 激活sheet
    wb1 = wb.worksheets[sheet]
    for i in range(parse.nObj):
        wb1.cell(line, column+i, str(targ_i[i]))  # 对应14行B
    wb.save("result.xlsx")  # 保存
def Read_cell(path,lines,col): #lines是一个range()
    # 导入需要读取Excel表格的路径
    data = xlrd.open_workbook(path)
    table = data.sheets()[0]
    # 创建一个空列表，存储Excel的数据
    tables = []
    # 将excel表格内容导入到tables列表中def
    # for rown in range(table.nrows):
    for rown in lines:
        tables.append(table.cell_value(rown, col))
        # tables.append(table.cell_value(rown, 1))
    return tables
def crowdingDistanceSort(Chrom, ranks):
    fits = target_pop(Chrom)
    nPop = Chrom.shape[0]
    nF = fits.shape[1]  # 目标个数
    dis = np.zeros(nPop)
    nR = ranks.max()  # 最大等级
    indices = np.arange(nPop)
    for r in range(nR + 1):
        rIdices = indices[ranks == r]  # 当前等级种群的索引
        rPops = Chrom[ranks == r]  # 当前等级的种群
        rFits = fits[ranks == r]  # 当前等级种群的适应度
        rSortIdices = np.argsort(rFits, axis=0)  # 对纵向排序的索引
        rSortFits = np.sort(rFits, axis=0)
        fMax = np.max(rFits, axis=0)
        fMin = np.min(rFits, axis=0)
        n = len(rIdices)
        for i in range(nF):
            orIdices = rIdices[rSortIdices[:, i]]  # 当前操作元素的原始位置
            j = 1
            while n > 2 and j < n - 1:
                if fMax[i] != fMin[i]:
                    dis[orIdices[j]] += (rSortFits[j + 1, i] - rSortFits[j - 1, i]) / \
                                        (fMax[i] - fMin[i])
                else:
                    dis[orIdices[j]] = np.inf
                j += 1
            dis[orIdices[0]] = np.inf
            dis[orIdices[n - 1]] = np.inf
    return dis
def optSelect(pops, fits, chrPops, chrFits):
    nPop, nChr = pops.shape
    nF = fits.shape[1]
    newPops = np.zeros((nPop, nChr), dtype=int)
    newFits = np.zeros((nPop, nF))
    # 合并父代种群和子代种群构成一个新种群
    MergePops = np.concatenate((pops, chrPops), axis=0)  # 拼接
    MergeFits = np.concatenate((fits, chrFits), axis=0)
    # MergeRanks = nonDominationSort(MergePops)  # 两个种群合并了，两倍大小的种群
    MergeRanks = nonDominationSort(MergeFits)  # 两个种群合并了，两倍大小的种群
    MergeDistances = crowdingDistanceSort(MergePops, MergeRanks)

    indices = np.arange(MergePops.shape[0])
    r = 0
    i = 0
    rIndices = indices[MergeRanks == r]  # 当前等级为r的索引
    while i + len(rIndices) <= nPop:
        newPops[i:i + len(rIndices)] = MergePops[rIndices]  # 精英策略，先把等级小的全都放进新种群中
        newFits[i:i + len(rIndices)] = MergeFits[rIndices]
        r += 1  # 当前等级+1
        i += len(rIndices)
        rIndices = indices[MergeRanks == r]  # 当前等级为r的索引

    if i < nPop:  # 还有一部分没放，就在当前rank按拥挤度才弄个大到小放进去填满
        rDistances = MergeDistances[rIndices]  # 当前等级个体的拥挤度
        rSortedIdx = np.argsort(rDistances)[::-1]  # 按照距离排序 由大到小
        surIndices = rIndices[rSortedIdx[:(nPop - i)]]
        newPops[i:] = MergePops[surIndices]
        newFits[i:] = MergeFits[surIndices]
    return (newPops, newFits)
def uniformpoint(N, M):  # N:pop_size(总生成点个数),M:目标个数3
    H1 = 1
    while (comb(H1 + M - 1, M - 1) <= N):  # H1是每个方向上的方向数
        H1 = H1 + 1
    H1 = H1 - 1
    # H1+M-1中M-1个插板排列组合，
    W = np.array(list(combinations(range(H1 + M - 1), M - 1))) - np.tile(np.array(list(range(M - 1))),
                                                                         (int(comb(H1 + M - 1, M - 1)), 1))
    W = (np.hstack((W, H1 + np.zeros((W.shape[0], 1)))) - np.hstack(
        (np.zeros((W.shape[0], 1)), W))) / H1  # hstack水平拼接
    # 向量过于稀疏的情况
    if H1 < M:
        H2 = 0
        while (comb(H1 + M - 1, M - 1) + comb(H2 + M - 1, M - 1) <= N):
            H2 = H2 + 1
        H2 = H2 - 1
        if H2 > 0:
            W2 = np.array(list(combinations(range(H2 + M - 1), M - 1))) - np.tile(np.array(list(range(M - 1))),
                                                                                  (int(comb(H2 + M - 1, M - 1)), 1))
            W2 = (np.hstack((W2, H2 + np.zeros((W2.shape[0], 1)))) - np.hstack(
                (np.zeros((W2.shape[0], 1)), W2))) / H2
            W2 = W2 / 2 + 1 / (2 * M)
            W = np.vstack((W, W2))  # 垂直拼接
    W[W < 1e-6] = 1e-6  # 所有元素大于0
    N = W.shape[0]
    return W, N  # 一开始N是popsize，后面因为要生成平均向量，所以，种群个数改成能够形成组合数Cmn的那个N
def Tchebycheff(x, lamb, z):
    temp = []
    targ = target(x)
    for i in range(len(targ)):
        temp.append(np.abs(targ[i] - z[i]) * lamb[i])
    return np.max(temp)
def nonDominationSort(PopFun):
    nPop,nF = PopFun.shape[0],PopFun.shape[1]
    ranks = np.zeros(nPop, dtype=np.int32)
    nPs = np.zeros(nPop)  # 每个个体p被支配解的个数
    sPs = []  # 每个个体支配的解的集合，把索引放进去
    for i in range(nPop):
        iSet = []  # 解i的支配解集
        for j in range(nPop):
            if i == j:
                continue
            isDom1 = PopFun[i] <= PopFun[j]
            isDom2 = PopFun[i] < PopFun[j]
            # 是否支配该解-> i支配j
            if sum(isDom1) == nF and sum(isDom2) >= 1:
                iSet.append(j)
                # 是否被支配-> i被j支配
            if sum(~isDom2) == nF and sum(~isDom1) >= 1:
                nPs[i] += 1
        sPs.append(iSet)  # 添加i支配的解的索引
    r = 0  # 当前等级为 0， 等级越低越好
    indices = np.arange(nPop)
    while sum(nPs == 0) != 0:
        rIdices = indices[nPs == 0]  # 当前被支配数为0的索引
        ranks[rIdices] = r
        for rIdx in rIdices:
            iSet = sPs[rIdx]
            nPs[iSet] -= 1
        nPs[rIdices] = -1  # 当前等级的被支配数设置为负数
        r += 1
    return ranks
def isDominates(s1, s2):  # x是否支配y
    return (s1 <= s2).all() and (s1 < s2).any()

def roulette(select_list):
    sum_val = sum(select_list)
    random_val = random.random()
    probability = 0  # 累计概率
    for i in range(len(select_list)):
        probability += select_list[i] / sum_val  # 加上该个体的选中概率
        if probability >= random_val:
            return i  # 返回被选中的下标
        else:
            continue
def TOPSIS(targ_pop, lamda):
    # 首先定义了一个4行4列的矩阵X，代表4个候选方案的4个属性值。然后，我们定义了一个长度为4的权重向量w，代表每个属性的权重。
    # 接着，我们定义了一个长度为4的布尔向量is_benefit，用于表示每个属性是否为效益属性。如果为效益属性，则为True；否则为False。在本例中，我们假设所有属性均为效益属性。
    # 接下来，我们调用topsis函数来计算TOPSIS得分。该函数首先对矩阵进行标准化处理，然后计算加权标准化矩阵。接着，我们计算正理想解和负
    # 标准化矩阵
    X = targ_pop
    w = lamda
    X_norm = np.zeros(X.shape)
    for i in range(X.shape[1]):
        X_norm[:, i] = (X[:, i] - np.min(X[:, i])) / (np.max(X[:, i]) - np.min(X[:, i]))
    # 计算加权标准化矩阵
    X_weighted = np.zeros(X.shape)
    for i in range(X.shape[1]):
        X_weighted[:, i] = X_norm[:, i] * w[i]
    # 计算正理想解和负理想解
    z_positive = np.zeros(X.shape[1])
    z_negative = np.zeros(X.shape[1])
    for i in range(X.shape[1]):
        z_positive[i] = np.max(X_weighted[:, i])
        z_negative[i] = np.min(X_weighted[:, i])
    # 计算距离正理想解和负理想解的欧几里得距离
    d_positive = np.sqrt(np.sum((X_weighted - z_positive)**2, axis=1))
    d_negative = np.sqrt(np.sum((X_weighted - z_negative)**2, axis=1))
    # 计算综合得分
    s = d_negative / (d_positive + d_negative)
    return s
def target_pop(Chrom):
    targ_pop = np.zeros((len(Chrom), parse.nObj))
    for i in range(len(Chrom)):
        PVal, fw, P, TWF, makespan, Pc, Lb= caltimecost_fatigue(Chrom[i])
        targ_pop[i, 0] = makespan
        targ_pop[i, 1] = Pc
        targ_pop[i, 2] = Lb
    return targ_pop
# SPEA2
def spea_cal_fitness(pop):  # 计算 Pt和 Et 适应度, F(i) = R(i) + D(i)
    objectives = []
    set = []
    length = len(pop)
    S = np.zeros(length)
    D = np.zeros(length)
    R = np.zeros(length)
    F = np.zeros(length)
    KNN = np.zeros(length)
    K = int(np.sqrt(length))
    for i in range(length):
        objectives.append(target(pop[i]))  #每个染色体的目标值
    # 计算 S 值
    for i in range(length):
        temp = []
        for j in range(length):
            if j != i:
                if isDominates(objectives[i],objectives[j]):
                    S[i] += 1  # i支配 j，np+1
                if isDominates(objectives[j],objectives[j]):
                    temp.append(j)  # j支配 i
        set.append(temp)
        # 计算 R 值
    for i in range(length):
        for j in range(len(set[i])):
            R[i] += S[set[i][j]]
    # 计算 D 值
    for i in range(length):
        distance = []
        for j in range(length):
            if j != i:
                distance.append(np.sqrt(np.square(objectives[i][0] - objectives[j][0]) + np.square(
                    objectives[i][1] - objectives[j][1])))
        distance = sorted(distance)
        KNN[i] = distance[K - 1]  # 其它个体与个体 i 的距离，升序排序，取第 K 个距离值
        D[i] = 1 / (KNN[i] + 2)
    # 计算 F 值
    for i in range(length):
        F[i] = D[i] + R[i]
    return F,D #所有种群个体的F和D

def update(F,D,mixpop,archive):  # 下一代 Archive
    juli = []
    shiyingzhi = []
    a = 0
    length = len(mixpop)
    for i in range(length):
        if F[i] < 1:
            juli.append([D[i], i])
            a = a + 1
        else:
            shiyingzhi.append([F[i], i])
    # 判断是否超出范围
    if a > parse.pop_size:  # 截断策略
        juli = sorted(juli)
        for i in range(parse.pop_size):
            archive[i] = mixpop[juli[i][1]]
    if a == parse.pop_size:  # 刚好填充
        for i in range(parse.pop_size):
            archive[i] = mixpop[juli[i][1]]
    if a < parse.pop_size:  # 适应值筛选
        shiyingzhi = sorted(shiyingzhi)
        for i in range(a):
            archive[i] = mixpop[juli[i][1]]
        for i in range(parse.pop_sizep - a):
            archive[i + a] = mixpop[shiyingzhi[i][1]]
    return archive
def selection(mixpop):  # 轮盘赌选择
    length = len(mixpop)
    pi = np.zeros(length)  # 个体的概率
    qi = np.zeros(length + 1)  # 个体的累积概率
    P = 0
    F,D = spea_cal_fitness(mixpop)  # 计算Archive的适应值
    for i in range(len(mixpop)):
        P += 1 / F[i]  # 取倒数，求累积适应度
    for i in range(len(mixpop)):
        pi[i] = (1 / F[i]) / P  # 个体遗传到下一代的概率
    for i in range(len(mixpop)):
        qi[0] = 0
        qi[i + 1] = np.sum(pi[0:i + 1])  # 累积概率
    # 生成新的 population
    Chrom = np.zeros((parse.pop_size, 3 * parse.W), dtype=int)
    for i in range(parse.pop_size):
        r = random.random()  # 生成随机数，
        for j in range(length):
            if r > qi[j] and r < qi[j + 1]:
                Chrom[i] = mixpop[j]
            j += 1
    return Chrom
#指标
def IGD(popfun,PF,Zmin, Zmax):
    num_PF = PF.shape[0]
    popfun = (popfun-Zmin[0])/(Zmax[0]-Zmin[0])
    PF = (PF-Zmin[0])/(Zmax[0]-Zmin[0])
    distances = np.zeros(num_PF)
    for i, p_true in enumerate(PF):
        distances[i] = np.min(np.linalg.norm(popfun - p_true, axis=1))
    igd = np.sum(distances) / num_PF
    return igd
def GD(popfun, PF,Zmin, Zmax):
    num_pop = popfun.shape[0]
    popfun = (popfun - Zmin[0]) / (Zmax[0] - Zmin[0])
    PF = (PF - Zmin[0]) / (Zmax[0] - Zmin[0])
    distances = np.zeros(num_pop)
    for i, pop in enumerate(popfun):
        distances[i] = np.min(np.linalg.norm(pop - PF, axis=1))
    gd = np.sqrt(np.sum(distances)) / num_pop
    return gd
def HV(pf,Zmin, Zmax):
    # 输入Pareto Front的列表
    # pf = [(1, 1, 1), (2, 2, 2), (3, 4, 3), (4, 3, 4), (5, 5, 5)],ref_point = [(6, 6, 6)]
    ref_point = np.array([1,1,1])
    pf = (pf - Zmin[0]) / (Zmax[0] - Zmin[0])
    # 定义超体积初始值
    hv = 0.0
    # 计算每个超体积贡献
    # ref_point_i = Zmin[0]
    for i in range(len(pf)):
        # 计算当前点到参考点的距离
        dist = (ref_point[0] - pf[i][0]) * (ref_point[1]-pf[i][1]) * (ref_point[2]-pf[i][2])
        # 如果距离为负数，则将其置为0
        if dist < 0:
            dist = 0
        # 累加超体积贡献
        hv += dist
    return hv
def Coverage(A, B):
    numB = 0  # B中被支配的个体
    for i in B:  # 对于集合B中的每一个个体
        for j in A:
            if isDominates(j, i):  # 至少被A中一个解支配
                numB += 1
                break
    ratio = numB / len(B)
    return ratio
#模型
def target(spring):
    targ = np.zeros(parse.nObj)
    PVal, fw,P,TWF, makespan, Pc, Lb = caltimecost_fatigue(spring)
    targ[0] = makespan
    targ[1] = Pc
    targ[2] = Lb
    return targ
def caltimecost_fatigue(C):  # 编码染色体个体C
    fw = np.zeros((30000, parse.Es))  # 记录每个人的疲劳度
    M = copy.deepcopy(C[parse.J * parse.S:parse.J * parse.S * 2])  # 机器编码
    E = copy.deepcopy(C[parse.J * parse.S * 2:parse.J * parse.S * 3])  # 人员编码
    TP = np.zeros((1, parse.J), dtype=int)  # 记录每个工件的上次完工时刻
    PVal = np.zeros((2, parse.W), dtype=int)  # 按染色体顺序记录每道工序加工所需时间、完成时刻
    TMF = np.zeros((3, parse.JmNumber), dtype=int)  # 记录机器上次完工时刻、累计总加工时间、上次加工工序
    TWF = np.zeros((4, parse.Es), dtype=int)  # 记录人员上次完工时间、上次加工机器、上次加工工件、上次加工时间
    P = decode(C[:parse.W])
    Pc = 0
    for i in range(parse.W):
        val = int(P[i])
        a = int(val % 100)  # 工序
        b = int((val - a) / 100)  # 工件
        mi = int(M[i])  # 机器序号
        ei = int(E[i])  # 加工该机器的人员(编号1开始)
        mi_index = parse.JM[b - 1][a - 1].index(mi)
        base_t = parse.TM[b - 1][a - 1][mi_index]  # 对应的机器加工基准时间
        ei_index = parse.JW[mi-1].index(ei)
        Pr = parse.Pr[mi-1][ei_index]  # 工人被分配到该工件（工序）的偏好，适应度，能力
        TMval = TMF[0][mi - 1]  # 机器上次的完工时刻
        TMpro = TMF[2][mi - 1]  # 机器上次加工的工序
        TPval = TP[0][b - 1]  # 工件上次完工时刻
        TWval = TWF[0][ei - 1]  # 工人上次完工时刻
        MWval = TWF[1][ei - 1] # 工人上次加工机器
        u = Pr * parse.u[int(ei) - 1]  # 偏好越大，人员恢复越快
        lamda = 1/Pr * parse.lamda[int(ei) - 1]  #偏好越大，疲劳率越小

        # if MWval > 0:
        if TWval > max(TMval, TPval): #如果要等工人结束上次工序之后才能开始加工
            trans_t = parse.Trans_t[MWval-1,mi-1]
        else:
            trans_t = 0
        startval = max(TMval, TPval, TWval)

        # 人员在开始前是休息的
        for k in range(int(TWval), int(startval)):
            fw[k][int(ei) - 1] = fw[k - 1][int(ei) - 1] * (math.e ** (-u))

        start_1 = startval  #没开始运输的时刻
        # 运输过程
        startval += trans_t
        for k in range(int(start_1), int(startval)):
            fw[k][int(ei) - 1] = fw[k - 1][int(ei) - 1]
        # for k in range(int(startval)-int(TWval)+1):
        #     fw[TWval+k][int(ei) - 1] = fw[int(TWval)][int(ei) - 1]

        # 判断人员是否需要休息
        ##计算空闲时间的疲劳度值变化
        start_2 = startval
        while fw[int(startval)][int(ei) - 1] >= parse.fmax:  # 如果人员在开始前一秒还是疲劳度过高的话就休息gap
            startval += parse.gap
            for k in range(int(start_2), int(startval)):
                fw[k][int(ei) - 1] = fw[k - 1][int(ei) - 1] * (math.e ** (-u))
        # akk = math.e ** (-u)
        # for k in range(int(startval)-int(rest_begin)+1):
        #     fw[rest_begin+k][int(ei) - 1] = math.pow(akk, k) * fw[rest_begin][int(ei) - 1]

        # if TMpro != a:  # 上次加工的工序和现在不同
        #     ChanTime = parse.TM[0][0][0]/10  # 换模时间固定
        # else:
        #     ChanTime = 0
        # t += ChanTime  # 加工时间含换模
        # t = int(t)

        ##计算加工工件时的疲劳度变化
        endval = int(startval) + base_t
        for j in range(int(startval), endval):
            fw[j][int(ei) - 1] = 1 - (1 - fw[j - 1][int(ei) - 1]) * math.e ** (-lamda)
        # att = math.e ** (-lamda)
        # btt = 1 - math.e ** (-lamda)
        # for k in range(t+1):
        #     fw[startval+k][int(ei) - 1] = math.pow(att, k) * fw[int(startval)][int(ei) - 1] + btt * (1-math.pow(att,k)/(1-att))
        ##修正一下时间（机加工上料下料时间、手工作业时间都会有影响）
        re_f = np.mean(fw[int(startval): endval,int(ei) - 1])
        if re_f < parse.f_1:
            re_t = 0
        elif re_f < parse.f_2:
            re_t = int(0.1 * base_t)
        else:
            re_t = int(0.2 * base_t)
        for j in range(endval, endval + re_t):
            fw[j][int(ei) - 1] = 1 - (1 - fw[j - 1][int(ei) - 1]) * math.e ** (-lamda)
        # for k in range(re_t+1):
        #     fw[startval+t+k][int(ei) - 1] = math.pow(att,k) * fw[int(startval+re_t)][int(ei) - 1] + btt * (1-math.pow(att,k)/(1-att))
        endval += re_t #结束时刻
        t = base_t + re_t

        # TMF[3][mi - 1] = t  # 加工时间不含换模*用于计算下一次换模时间
        PVal[0][i] = t  # 工件工序加工时长
        PVal[1][i] = endval  # 工件工序加工完成时刻
        TMF[0][mi - 1] = endval  # 机器加工完成时刻
        TMF[1][mi - 1] += t  # 机器加工总时长
        TMF[2][mi - 1] = a  # 机器加工工序
        TP[0][b - 1] = endval  # 工件加工完成时刻
        TWF[0][ei - 1] = endval  # 人员加工完成时刻
        TWF[1][ei - 1] = mi  # 人员加工机器
        TWF[2][ei - 1] = b # 人员加工工件
        TWF[3][ei - 1] = t # 人员加工时间

        # 人员没有任务后疲劳度的变化情况
        if ei not in E[i + 1:]:
            for k in range(endval, endval + 500):
                fw[k][int(ei) - 1] = fw[k - 1][int(ei) - 1] * math.e ** (-u)

        Pc += parse.m_cost[mi-1]*t+parse.w_cost[ei-1]*t

    # 计算三个目标值
    makespan = max(PVal[1][:])
    temp = np.zeros(parse.Es)
    for i in range(parse.Es):  # 每个员工加工过程中的疲劳度总和
        temp[i] = np.sum(fw[:makespan, i])
    Fmean = np.sum(temp)/parse.Es
    # SI = math.sqrt(np.sum((TWF[3] - np.mean(TWF[3]))** 2) / parse.Es)
    # # 每个员工的疲劳度与均值插值平方求和再开方=每个员工满意度part1
    Lb = math.sqrt(np.sum((temp-Fmean)**2)/parse.Es)

    # # 每个员工的分配工作满意度=每个员工满意度part2
    #  = parse.alpha[0]*Eq_f + parse.alpha[1]*Eq_s
    # MachineLoad = np.sum(TMF[1])  # 所有机器的总完成时间
    return PVal, fw,P,TWF, makespan,Pc,Lb

# def penalty_based_boundary(x, lamb, z):
#     temp = []
#     targ = target(x)
#     targ = (targ - targ.min(axis=0)) / (targ.max(axis=0) - targ.min(axis=0)) ## 归一化
#     for i in range(len(targ)):
#         d1 = np.linalg.norm(np.dot(z - targ[i],lamb[i]))/np.linalg.norm(lamb[i])
#         d2 = np.linalg.norm(targ[i] - (z - d1*lamb[i]))
#         return d1 + parse.Hf * d2

def penalty_based_boundary(x, lamb, z):
    targ = target(x)
    d1 = np.linalg.norm(np.dot(targ - z, lamb))/np.linalg.norm(lamb)
    d2 = np.linalg.norm(targ - (z + d1*lamb/np.linalg.norm(lamb)))
    return d1 + parse.Hf * d2

#画图
def plot_figure_worker(string): #machine
    # ---------------甘特图-------------
    PVal, fw,P,TWF, makespan, Pc, Lb = caltimecost_fatigue(string)
    # PVal, fw, P, makespan, FwTotal, MachineLoad, Chrom, z, popfun, score = self.abtain_chrom()
    print("makespan", makespan)
    print('Pc', Pc)
    print('Lb', Lb)
    cst = PVal[1][:] - PVal[0][:]  # 每道工序开始加工时刻
    cpt = PVal[0][:]  # 每道工序的加工时长
    cft = PVal[1][:]  # 每道工序的加工完成时刻
    cmn = np.zeros(parse.W, dtype=int)  # 对应的机器号
    cwn = np.zeros(parse.W, dtype=int)  # 对应人员编号
    M = copy.deepcopy(string[parse.W:parse.W * 2])  # 机器编码
    E = copy.deepcopy(string[parse.W* 2:parse.W * 3])  # 人员编码
    for i in range(parse.W):
        val = int(P[i])
        a = int(val % 100)  # 工序
        b = int((val - a) / 100)  # 工件
        mi = int(M[i])  # 机器序号
        ei = int(E[i])  # 人员序号
        cmn[i] = mi
        cwn[i] = ei
    osc = np.tile(string[0:parse.W], 1)

    # ---------------甘特图-------------
    plt.figure()
    for i in range(parse.W):
        if cft[i] != 0:
            plt.barh(y=cwn[i], width=cft[i] - cst[i], height=0.8, left=cst[i],
                     color=COLORS[osc[i] % LEN_COLORS], alpha=0.8, edgecolor="black")
            sf = r"$_{%s}$" % cst[i], r"$_{%s}$" % cft[i]
            # x = cst[i], cft[i]
            # for j, k in enumerate(sf):  # 时间刻度
            #     plt.text(x=x[j], y=cmn[i], s=k,
            #              rotation="horizontal", va="top", ha="center")
            text = r"${%s}$" % (osc[i])  # 工件编号
            # text1 = r"${w%s}$" % (cwn[i])
            plt.text(x=cst[i] + 0.5 * cpt[i], y=cwn[i], s=text, c="black",
                     rotation="horizontal", va="center", ha="center")
            # plt.text(x=cst[i] + 0.8 * cpt[i], y=cmn[i], s=text1, c="black",
            #          rotation="horizontal", va="center", ha="center")
    plt.ylabel(r"Worker", fontsize=12, fontproperties="Arial")
    plt.xlabel(r"Makespan", fontsize=12, fontproperties="Arial")
    plt.title(r"Gantt Chart", fontsize=14, fontproperties="Arial")
    plt.show()
def plot_figure_machine(string):
    # ---------------甘特图-------------
    start = time.time()
    PVal, fw,P,TWF, makespan, Pc, Lb = caltimecost_fatigue(string)
    # PVal, fw, P, makespan, FwTotal, MachineLoad, Chrom, z, popfun, score = self.abtain_chrom()
    end = time.time()
    print("calculation time：%2f second" % (end - start))
    print("makespan", makespan)
    print('Pc', Pc)
    print('Lb', Lb)
    cst = PVal[1][:] - PVal[0][:]  # 每道工序开始加工时刻
    cpt = PVal[0][:]  # 每道工序的加工时长
    cft = PVal[1][:]  # 每道工序的加工完成时刻
    cmn = np.zeros(parse.W, dtype=int)  # 对应的机器号
    cwn = np.zeros(parse.W, dtype=int)  # 对应人员编号
    M = copy.deepcopy(string[parse.W:parse.W * 2])  # 机器编码
    E = copy.deepcopy(string[parse.W* 2:parse.W * 3])  # 人员编码
    for i in range(parse.W):
        val = int(P[i])
        a = int(val % 100)  # 工序
        b = int((val - a) / 100)  # 工件
        mi = int(M[i])  # 机器序号
        ei = int(E[i])  # 人员序号
        cmn[i] = mi
        cwn[i] = ei
    osc = np.tile(string[0:parse.W], 1)

    # ---------------甘特图-------------
    plt.figure()
    for i in range(parse.W):
        if cft[i] != 0:
            plt.barh(y=cmn[i], width=cft[i] - cst[i], height=0.8, left=cst[i],
                     color=COLORS[osc[i] % LEN_COLORS], alpha=0.8, edgecolor="black")
            sf = r"$_{%s}$" % cst[i], r"$_{%s}$" % cft[i]
            x = cst[i], cft[i]
            # for j, k in enumerate(sf):  # 时间刻度
            #     plt.text(x=x[j], y=cmn[i], s=k,
            #              rotation="horizontal", va="top", ha="center")
            text = r"${%s}$" % (osc[i])  # 工件编号
            # text1 = r"${w%s}$" % (cwn[i])
            plt.text(x=cst[i] + 0.5 * cpt[i], y=cmn[i], s=text, c="black",
                     rotation="horizontal", va="center", ha="center")
            # plt.text(x=cst[i] + 0.8 * cpt[i], y=cmn[i], s=text1, c="black",
            #          rotation="horizontal", va="center", ha="center")
    plt.ylabel(r"Machine", fontsize=12, fontproperties="Arial")
    plt.xlabel(r"Makespan", fontsize=12, fontproperties="Arial")
    plt.title(r"Gantt Chart", fontsize=14, fontproperties="Arial")
    plt.show()
def ScatterPlot(A):
    fig = plt.figure()
    ax = Axes3D(fig)
    ax.scatter(A[:, 0], A[:, 1], A[:, 2], marker="o", color="y")
    ax.set_xlabel('MakeSpan')
    ax.set_ylabel('SI')
    ax.set_zlabel('Fmean')  # 给三个坐标轴注明坐标名称
    plt.title("Three objectives ScatterPlot")
    plt.grid(True)
    plt.legend()
    plt.show()
def plot_fatig(string):
    plt.figure()
    PVal, fw, P, TWF, makespan, Pc, Lb = caltimecost_fatigue(string)
    fw = np.round(fw, 4)
    fw = fw[:int(makespan) + 1, :]
    x = range(0, int(makespan) + 1)
    y1 = fw[:, 0]
    y2 = fw[:, 1]
    y3 = fw[:, 2]
    y4 = fw[:, 3]
    y5 = fw[:, 4]
    y6 = fw[:, 5]
    plt.plot(x, y1, label='Worker 1', linewidth=2, color='darkorange')
    plt.plot(x, y2, label='Worker 2', linewidth=2, color='pink')
    plt.plot(x, y3, label='Worker 3', linewidth=2, color='darkseagreen')
    plt.plot(x, y4, label='Worker 4', linewidth=2, color='lightgreen')
    plt.plot(x, y5, label='Worker 5', linewidth=2, color='lightskyblue')
    plt.plot(x, y6, label='Worker 6', linewidth=2, color='violet')
    plt.xlabel('Time', fontsize=14, fontproperties="Arial")
    plt.ylabel('Fatigue', fontsize=14, fontproperties="Arial")
    plt.title('Fatigue Curve', fontsize=14, fontproperties="Arial")
    plt.legend()
    plt.show()
def Compare_ScatterPlot(A,B,C,D):
    fig = plt.figure()
    ax = Axes3D(fig)
    type1 = ax.scatter(A[:, 0], A[:, 1], A[:, 2], marker="o", color="y")
    type2 = ax.scatter(B[:, 0], B[:, 1], B[:, 2], marker="o", color="b")
    type3 = ax.scatter(C[:, 0], C[:, 1], C[:, 2], marker="o", color="g")
    type4 = ax.scatter(D[:, 0], D[:, 1], D[:, 2], marker="o", color="r")
    ax.legend((type1, type2,type3,type4), ("Algorithm Without perturbation operators","Algorithm Whithout Metropolis guidelines","TSEA","Original MOEA/D",), loc=0)
    ax.set_xlabel('Makespan')
    ax.set_ylabel('Production Cost')
    ax.set_zlabel('Load Balance')  # 给三个坐标轴注明坐标名称
    plt.title("Three objectives ScatterPlot")
    plt.grid(True)
    plt.legend()
    plt.show()
def plot_box_diagram(A):
    fig = plt.figure()
    sp = fig.add_axes([0.1,0.1,0.8,0.8])
    sp.set_title("Comparison of Algorithm Improvements")
    matplotlib.rc("font", family='Times New Roman')
    plt.ylabel('time(s)', fontsize=18)

    # # 绘图
    plt.boxplot(A)
    # plt.boxplot(A, widths=0.4,
    #            patch_artist=True, showfliers=False,
    #            boxprops={'facecolor': 'skyblue', 'linewidth': 0.8, 'edgecolor': 'black'}, meanline=True,
    #            meanprops={'color': 'red', 'linewidth': 3})
    # 设置轴坐标值刻度的标签
    # plt.set_xticklabels(["Original MOEA/D", "Whithout Metropolis", "Without perturbation", "TSEA"], fontsize=14)
    plt.show()

def zhuzhuangtu():
    size = 8
    x = np.arange(size)

    # 有a/b两种类型的数据，n设置为2
    total_width, n = 0.6, 2
    # 每种类型的柱状图宽度
    width = total_width / n
    list1 = [88.59, 86.23, 87.79, 84.35, 89.97, 88.36, 85.51, 89.99]
    list2 = [79.92, 76.53, 79.32, 76.17, 79.78, 80.92, 77.51, 81.23]
    # 重新设置x轴的坐标
    x = x - (total_width - width) / 2
    print(x)
    plt.rcParams['font.serif'] = ['Times New Roman']
    # 画柱状图
    plt.bar(x, list1, width=width, label="Coarse", color='#0066cc')
    plt.bar(x + width, list2, width=width, label="Fine", color='#9ACD32')
    # plt.bar(x + 2*width, c, width=width, label="c")
    plt.xticks(np.arange(8), ('a', 'b', 'c', 'd', 'e', 'f', 'g', 'ours'))
    # 显示图例
    # plt.figure(dpi=300,figsize=(24,24))
    plt.legend(loc='lower right', prop={"family": "Times New Roman"})
    plt.xlabel("Comparision   Experiments", fontname="Times New Roman")
    plt.ylabel("Dice  Score", fontname="Times New Roman")
    plt.savefig('plot123_2.png', dpi=500)
    # 显示柱状图
    plt.show()


#算法优化
def initialize():
    Chrom = np.zeros((parse.pop_size, 3 * parse.W), dtype=int)  # 三层编码【工序】【机器】【人员】
    for j in range(parse.pop_size):  # 第j条染色体
        # 工件序列OS
        OS = np.zeros(parse.W, dtype=int)
        num = 0
        for job in range(parse.J):
            for k in range(parse.S):
                OS[num] = job + 1  # OS=[ 1  1  1  2  2  2  3  3  3 ... 12 12 12]
                num += 1
        np.random.shuffle(OS)  # 打乱
        # 机器序列MS,人员序列WS
        MS = np.zeros(parse.W, dtype=int)
        WS = np.zeros(parse.W, dtype=int)
        P = decode(OS)
        for k in range(parse.W):
            val = int(P[k])
            a = int(val % 100)  # 工序
            b = int((val - a) / 100)  # 工件
            tmp_m = parse.JM[b-1][a-1]  ##可选机器
            MS[k] = np.random.choice(tmp_m)  ##随机选择一个机器
            tmp_w = parse.JW[MS[k] - 1]  ##机器对应的可选人员
            WS[k] = np.random.choice(tmp_w)  ##随机选择一个人员
        Chrom[j][:parse.W] = OS
        Chrom[j][parse.W:2 * parse.W] = MS
        Chrom[j][2 * parse.W:3 * parse.W] = WS
    return Chrom  # 初始化Chrom为一个种群
def look_neighbor(lamda):
    B = []
    for i in range(len(lamda)):
        temp = []
        for j in range(len(lamda)):
            distance = np.sqrt((lamda[i][0] - lamda[j][0]) ** 2 +
                               (lamda[i][1] - lamda[j][1]) ** 2 + (lamda[i][2] - lamda[j][2]) ** 2)
            temp.append(distance)
        l = np.argsort(temp)
        B.append(l[:parse.T])
    return B
def crossover(s1,s2):
    # 两点交叉
    if random.random()<parse.XOVR:
        pos1,pos2 = sorted(random.sample(range(1,parse.W-2), 2))
        offspring = np.zeros(3 * parse.W, dtype=int)
        offspring[:pos1] = s1[:pos1]
        offspring[parse.W:parse.W + pos1] = s1[parse.W:parse.W + pos1]
        offspring[parse.W*2:parse.W*2 + pos1] = s1[parse.W*2:parse.W*2 + pos1]
        offspring[pos1:pos2] = s2[pos1:pos2]
        offspring[parse.W + pos1:parse.W + pos2] = s2[parse.W + pos1:parse.W + pos2]
        offspring[parse.W*2 + pos1:parse.W*2 + pos2] = s2[parse.W*2 + pos1:parse.W*2 + pos2]
        offspring[pos2:parse.W] = s1[pos2:parse.W]
        offspring[pos2+parse.W:parse.W*2] = s1[pos2+parse.W:parse.W*2]
        offspring[parse.W*2 + pos2:] = s1[parse.W*2 + pos2:]

        #检查可行性，修正染色体
        OS = offspring[:parse.W]
        MS = offspring[parse.W:parse.W*2]
        WS = offspring[parse.W*2:parse.W*3]
        temp = np.zeros(parse.J, dtype=int)  # 子代中每个工件已经有几道工序了
        for k in range(parse.W):
            while temp[OS[k] - 1] >= parse.S:  # 对应的这个基因在子代中到达最大工序数
                OS[k] = random.randint(1,parse.J)
            temp[OS[k] - 1] += 1
        P = decode(OS)
        for k in range(parse.W):
            val = int(P[k])
            a = int(val % 100)  # 工序
            b = int((val - a) / 100)  # 工件
            tmp_m = parse.JM[b - 1][a - 1]  ##可选机器
            if MS[k] not in tmp_m:
                MS[k] = np.random.choice(tmp_m)  ##随机选择一个机器
            tmp_w = parse.JW[MS[k] - 1]  ##机器对应的可选人员
            if WS[k] not in tmp_w:
                WS[k] = np.random.choice(tmp_w)  ##随机选择一个人员
        offspring[:parse.W] = OS
        offspring[parse.W:2 * parse.W] = MS
        offspring[2 * parse.W:3 * parse.W] = WS
        # caltimecost_fatigue(offspring)
    else:
        offspring = random.choice((s1,s2))
    return offspring
def mutation(offspring):
    P = decode(offspring[:parse.W]) #解码后的
    choice_pos_1 = np.random.choice(parse.W,14,replace=False) #选14个位置变异
    if random.random()<parse.MUTR:
        for j in range(len(choice_pos_1)):
            pos = choice_pos_1[j]
            val = int(P[pos])
            a = int(val % 100)  # 工序
            b = int((val - a) / 100)  # 工件
            mi_list = parse.JM[b-1][a-1] #机器列表
            offspring[parse.W+pos] = np.random.choice(mi_list)  ##随机选择一个机器
    choice_pos_2 = np.random.choice(parse.W, 14, replace=False)  # 选14个位置变异
    if random.random()<parse.MUTR:
        for j in range(len(choice_pos_2)):
            pos = choice_pos_2[j]
            mi = offspring[parse.W + pos]
            ei_list = parse.JW[mi - 1]  ##机器对应的可选人员
            offspring[2 * parse.W + pos] = np.random.choice(ei_list)  ##随机选择一个人员
    else:
        pass
    return offspring
class local_search():
    # def LS1(self,s):
    #     # 倒序变换
    #     W = parse.W
    #     parentspring = copy.deepcopy(s)
    #     length = len(parentspring) // 3
    #     loca1 = random.randint(0, length - 1)
    #     loca2 = random.randint(0, length - 1)
    #     offspring = copy.deepcopy(parentspring)
    #     while loca2 == loca1:  # 如果相等就重新生成
    #         loca2 = random.randint(0, length - 1)
    #     if loca2 < loca1:  # 保证loca1<loca2
    #         temp = loca2
    #         loca2 = loca1
    #         loca1 = temp
    #     for i in range(loca1, loca2 + 1):
    #         offspring[i] = parentspring[loca2 - i + loca1]
    #         offspring[i + W] = parentspring[loca2 - i + loca1 + W]
    #         offspring[i + W * 2] = parentspring[loca2 - i + loca1 + W * 2]
    #     return offspring
    # 要检查可行解
    def LS1(self, s):
        # 基因片段前后交换(这两个如果工件不一样的话),往前挪一格
        pos = random.randint(0,parse.W-2)
        if s[pos] != s[pos+1]:
            temp_1,temp_2,temp_3 = s[pos], s[pos+parse.W],s[pos+parse.W*2]
            s[pos], s[pos + parse.W], s[pos + parse.W * 2] = s[pos+1], s[pos+parse.W+1],s[pos+parse.W*2+1]
            s[pos + 1], s[pos + parse.W + 1], s[pos + parse.W * 2 + 1] = temp_1,temp_2,temp_3
        return s

    def LS2(self,s):
        # 操作是确定的
        # 同一设备组的机器人员两两交换
        pos = random.randint(0,parse.W-1)
        mi = s[pos + parse.W] #机器序号
        pos_m = random.randint(0, parse.W - 1)
        m_i = s[pos_m + parse.W]
        if mi < 4:
            while m_i >= 4:
                pos_m = random.randint(0, parse.W - 1)
                m_i = s[pos_m + parse.W]
        elif mi in (4,5):
            while not m_i in (4,5):
                pos_m = random.randint(0, parse.W - 1)
                m_i = s[pos_m + parse.W]
        elif mi in (6,7):
            while not m_i in (6, 7):
                pos_m = random.randint(0, parse.W - 1)
                m_i = s[pos_m + parse.W]
        elif mi == 8:
            while not m_i == 8:
                pos_m = random.randint(0, parse.W - 1)
                m_i = s[pos_m + parse.W]
        elif mi == 9:
            while not m_i == 9:
                pos_m = random.randint(0, parse.W - 1)
                m_i = s[pos_m + parse.W]

        temp_5 = mi
        s[pos + parse.W] = s[pos_m + parse.W]
        s[pos_m + parse.W] = temp_5
        temp_4 = s[pos + 2 * parse.W]
        s[pos + 2 * parse.W] = s[pos_m + 2 * parse.W]
        s[pos_m + 2 * parse.W] = temp_4
        return s


    # def LS2(self,s):
    #     # 任选两个位置交换
    #     W = parse.W
    #     loca1,loca2 = np.random.choice(len(s)//3,2,replace=False)
    #     new_s = copy.deepcopy(s)
    #     temp = np.zeros(3,dtype=int)
    #     temp[0] = s[loca1]
    #     temp[1] = s[loca1+W]
    #     temp[2] = s[loca1+W*2]
    #     new_s[loca1] = s[loca2]
    #     new_s[loca1+W] = s[loca2+W]
    #     new_s[loca1+W*2] = s[loca2+W*2]
    #     new_s[loca2] = temp[0]
    #     new_s[loca2 + W] = temp[1]
    #     new_s[loca2 + W * 2] = temp[2]
    #     return new_s
    def LS3(self,s):
        # 任选一个位置变机器
        pos = random.randint(0,parse.W-1)
        mi = s[pos + parse.W]  #机器编号
        if mi < 4:
            s[pos + parse.W] = random.randint(1,3)
        elif mi == 4:
            s[pos + parse.W] = 5
        elif mi == 5:
            s[pos + parse.W] = 4
        elif mi == 6:
            s[pos + parse.W] = 7
        elif mi == 7:
            s[pos + parse.W] = 6
        else:
            pass
        # W = parse.W
        # JmNumber = parse.JmNumber
        # loca = np.random.randint(0,len(s) // 3-1)
        # new_s = copy.deepcopy(s)
        # new_s[loca+W] = random.randint(1, JmNumber)
        return s

    def LS4(self,s):
        # 任选一个位置变人员
        # W = parse.W
        # JW = parse.JW
        # loca = np.random.randint(0,len(s) // 3-1)
        # new_s = copy.deepcopy(s)
        # mi = new_s[loca + W]
        # wList = JW[int(mi) - 1]  # 对应员工列表
        # new_s[loca + W*2] = np.random.choice(wList)
        pos = random.randint(0,parse.W-1)
        mi = s[pos + parse.W]
        s[pos + 2 * parse.W] = np.random.choice(parse.JW[mi-1])
        return s
    # def LS5(self,s):
    #     # 在最后一道工序替换最小工作负载机器
    #     J = parse.J
    #     W = parse.W
    #     TMF = parse.TMF
    #     new_s = copy.deepcopy(s)
    #     job = random.randint(0,J-1)
    #     pos = 0
    #     for pos_i in range(len(s)//3-1,-1,-1):
    #         if s[pos_i] == job:
    #             pos = pos_i
    #             break
    #     new_s[pos + W] = roulette(TMF) + 1
    #     return new_s
    # def LS6(self, s, fatigue):
    #     # 在最后一道工序替换最小疲劳度人员
    #     J = parse.J
    #     W = parse.W
    #     new_s = copy.deepcopy(s)
    #     job = random.randint(0, J - 1)
    #     pos = 0
    #     for pos_i in range(len(s) // 3 - 1, -1, -1):
    #         if s[pos_i] == job:
    #             pos = pos_i
    #             break
    #     # new_s[pos + W*2] = np.argmin(fatigue) + 1
    #     new_s[pos + W * 2] = roulette(fatigue) + 1
    #     return new_s


    # def LS5(self,s):
    #     # 在同一设备组选择人员累计疲劳度最小的那个人
    #     # 先选定一个位置,然后看机器属于哪个设备组,那个设备组可以选一个之前最小疲劳度的人

    # def LS6(self,s):
    #     # 在同一设备组选择耗时最小的机器
    #     # 先选定一个位置,然后看机器属于哪个设备组,那个设备组可以选一个之前最小疲劳度的人
# def SPT():
def cross_mutation(s1,s2):
    if random.random()<parse.XOVR:
        temp = np.zeros(parse.J, dtype=int)  # 子代中每个工件已经有几道工序了
        offspring = np.zeros(3 * parse.W, dtype=int)
        at1 = 0  # parent1指针
        at2 = 0  # parent2指针
        at = True  # 从哪个parent复制
        for i in range(len(offspring) // 3):
            while (offspring[i] == 0):  # 直到被赋值
                if at:  # 从parent1取基因
                    if temp[s1[at1] - 1] < parse.S:  # parent1对应的这个基因在子代中还没到达最大工序数
                        offspring[i] = s1[at1]  # 赋值
                        offspring[i + parse.W] = s1[at1 + parse.W]
                        offspring[i + parse.W * 2] = s1[at2 + parse.W * 2]
                    at1 += 1  # 不管是否赋值，at1指针向后一格
                    at = not at  # 逻辑取反，下次从parent2取
                else:  # 从parent2取基因
                    if temp[s2[at2] - 1] < parse.S:
                        offspring[i] = s2[at2]
                        offspring[i + parse.W] = s2[at2 + parse.W]
                        offspring[i + parse.W * 2] = s2[at2 + parse.W * 2]
                    at2 += 1
                    at = not at  # 逻辑取反
            temp[offspring[i] - 1] += 1
    else:
        offspring = random.choice((s1,s2))

    # 检查可行性，修正染色体
    OS = offspring[:parse.W]
    MS = offspring[parse.W:parse.W * 2]
    WS = offspring[parse.W * 2:parse.W * 3]
    temp = np.zeros(parse.J, dtype=int)  # 子代中每个工件已经有几道工序了
    for k in range(parse.W):
        while temp[OS[k] - 1] >= parse.S:  # 对应的这个基因在子代中到达最大工序数
            OS[k] = random.randint(1, parse.J)
        temp[OS[k] - 1] += 1
    P = decode(OS)
    for k in range(parse.W):
        val = int(P[k])
        a = int(val % 100)  # 工序
        b = int((val - a) / 100)  # 工件
        tmp_m = parse.JM[b - 1][a - 1]  ##可选机器
        if MS[k] not in tmp_m:
            MS[k] = np.random.choice(tmp_m)  ##随机选择一个机器
        tmp_w = parse.JW[MS[k] - 1]  ##机器对应的可选人员
        if WS[k] not in tmp_w:
            WS[k] = np.random.choice(tmp_w)  ##随机选择一个人员
    offspring[:parse.W] = OS
    offspring[parse.W:2 * parse.W] = MS
    offspring[2 * parse.W:3 * parse.W] = WS
    # caltimecost_fatigue(offspring)

    choice_pos_1 = np.random.choice(parse.W,14,replace=False) #选14个位置变异
    if random.random()<parse.MUTR:
        for j in range(len(choice_pos_1)):
            pos = choice_pos_1[j]
            val = int(P[pos])
            a = int(val % 100)  # 工序
            b = int((val - a) / 100)  # 工件
            mi_list = parse.JM[b-1][a-1] #机器列表
            offspring[parse.W+pos] = np.random.choice(mi_list)  ##随机选择一个机器
    choice_pos_2 = np.random.choice(parse.W, 14, replace=False)  # 选14个位置变异
    if random.random()<parse.MUTR:
        for j in range(len(choice_pos_2)):
            pos = choice_pos_2[j]
            mi = offspring[parse.W + pos]
            ei_list = parse.JW[mi - 1]  ##机器对应的可选人员
            offspring[2 * parse.W + pos] = np.random.choice(ei_list)  ##随机选择一个人员
    else:
        pass
    return offspring

